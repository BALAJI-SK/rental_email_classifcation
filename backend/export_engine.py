import json
import os
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXPORTS_DIR = os.path.join(os.path.dirname(__file__), "..", "exports")
os.makedirs(EXPORTS_DIR, exist_ok=True)

URGENCY_FILLS = {
    "critical": PatternFill("solid", fgColor="FECACA"),
    "high": PatternFill("solid", fgColor="FED7AA"),
    "medium": PatternFill("solid", fgColor="FEF08A"),
    "low": PatternFill("solid", fgColor="BBF7D0"),
}
HEADER_FILL = PatternFill("solid", fgColor="1E293B")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _header_row(ws, columns: list[str]):
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _auto_size(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


def _urgency_fill(ws, row_idx: int, urgency: str, col_count: int):
    fill = URGENCY_FILLS.get((urgency or "").lower())
    if fill:
        for col in range(1, col_count + 1):
            ws.cell(row=row_idx, column=col).fill = fill


def _ts() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


async def export_to_excel(db, export_type: str, filters: dict = None) -> str:
    filters = filters or {}
    prop_filter = filters.get("property_id")

    wb = Workbook()
    ws = wb.active

    filename = f"lette_{export_type}_{_ts()}.xlsx"
    filepath = os.path.join(EXPORTS_DIR, filename)
    row_count = 0

    if export_type == "open_issues":
        ws.title = "Open Issues"
        cols = ["Thread ID", "Subject", "Property", "Unit", "Category", "Urgency",
                "Days Open", "Contact", "Contact Phone", "Contact Email", "Summary",
                "Recommended Action", "Status"]
        _header_row(ws, cols)

        conditions = ["t.status IN ('open','in_progress','escalated')"]
        params = []
        if prop_filter:
            conditions.append("t.property_id = ?")
            params.append(prop_filter)

        async with db.execute(f"""
            SELECT t.id, t.subject, t.property_name, c.unit, t.category,
                   t.urgency_level, t.days_open, c.name, c.phone, c.email,
                   t.ai_summary, t.recommended_actions, t.status
            FROM threads t
            LEFT JOIN contacts c ON t.primary_contact_id = c.id
            WHERE {' AND '.join(conditions)}
            ORDER BY t.urgency_score DESC NULLS LAST
        """, params) as cursor:
            rows = await cursor.fetchall()

        for r_idx, row in enumerate(rows, 2):
            actions = []
            try:
                actions = json.loads(row[11] or "[]")
            except Exception:
                pass
            top_action = actions[0]["action"] if actions else ""
            data = [row[0], row[1], row[2], row[3], row[4], row[5],
                    row[6], row[7], row[8], row[9], row[10], top_action, row[12]]
            for c_idx, val in enumerate(data, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
            _urgency_fill(ws, r_idx, row[5], len(cols))
        row_count = len(rows)

    elif export_type == "tenant_contacts":
        ws.title = "Tenant Contacts"
        cols = ["Name", "Email", "Phone", "Unit", "Property", "Lease End",
                "Open Issues", "Total Messages", "Sentiment", "Last Contact"]
        _header_row(ws, cols)

        conditions = ["c.type = 'tenant'"]  # already qualified
        params = []
        if prop_filter:
            conditions.append("c.property_id = ?")
            params.append(prop_filter)

        async with db.execute(f"""
            SELECT c.name, c.email, c.phone, c.unit, p.name, c.lease_end,
                   c.total_threads, c.total_messages, c.sentiment_avg, c.last_seen
            FROM contacts c
            LEFT JOIN properties p ON c.property_id = p.id
            WHERE {' AND '.join(conditions)}
            ORDER BY p.name, c.name
        """, params) as cursor:
            rows = await cursor.fetchall()

        for r_idx, row in enumerate(rows, 2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        row_count = len(rows)

    elif export_type == "overdue_responses":
        ws.title = "Overdue Responses"
        cols = ["Thread ID", "Subject", "Tenant", "Property", "Days Waiting",
                "Follow-ups", "Urgency", "Recommended Action"]
        _header_row(ws, cols)

        conditions = ["t.status = 'open'", "t.days_open > 3"]
        params = []
        if prop_filter:
            conditions.append("t.property_id = ?")
            params.append(prop_filter)

        async with db.execute(f"""
            SELECT t.id, t.subject, c.name, t.property_name, t.days_open,
                   t.follow_up_count, t.urgency_level, t.recommended_actions
            FROM threads t
            LEFT JOIN contacts c ON t.primary_contact_id = c.id
            WHERE {' AND '.join(conditions)}
            ORDER BY t.days_open DESC
        """, params) as cursor:
            rows = await cursor.fetchall()

        for r_idx, row in enumerate(rows, 2):
            actions = []
            try:
                actions = json.loads(row[7] or "[]")
            except Exception:
                pass
            top_action = actions[0]["action"] if actions else ""
            data = [row[0], row[1], row[2], row[3], row[4], row[5], row[6], top_action]
            for c_idx, val in enumerate(data, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
            _urgency_fill(ws, r_idx, row[6], len(cols))
        row_count = len(rows)

    elif export_type == "property_report":
        ws.title = "Property Report"
        cols = ["Property", "Type", "Units", "Manager", "Total Threads",
                "Critical", "High", "Medium", "Low", "Unread", "Top Issue Category"]
        _header_row(ws, cols)

        async with db.execute("""
            SELECT p.id, p.name, p.type, p.units, p.manager,
                   COUNT(t.id) as total,
                   SUM(CASE WHEN t.urgency_level='critical' THEN 1 ELSE 0 END) as crit,
                   SUM(CASE WHEN t.urgency_level='high' THEN 1 ELSE 0 END) as high,
                   SUM(CASE WHEN t.urgency_level='medium' THEN 1 ELSE 0 END) as med,
                   SUM(CASE WHEN t.urgency_level='low' THEN 1 ELSE 0 END) as low,
                   SUM(CASE WHEN t.is_read=0 THEN 1 ELSE 0 END) as unread
            FROM properties p
            LEFT JOIN threads t ON t.property_id = p.id
            GROUP BY p.id
            ORDER BY p.name
        """) as cursor:
            rows = await cursor.fetchall()

        for r_idx, row in enumerate(rows, 2):
            prop_id = row[0]
            # Top category
            async with db.execute("""
                SELECT category, COUNT(*) as cnt FROM threads
                WHERE property_id=? AND category IS NOT NULL
                GROUP BY category ORDER BY cnt DESC LIMIT 1
            """, (prop_id,)) as c:
                top_cat = await c.fetchone()
            top_category = top_cat[0] if top_cat else "N/A"

            data = [row[1], row[2], row[3], row[4], row[5],
                    row[6], row[7], row[8], row[9], row[10], top_category]
            for c_idx, val in enumerate(data, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        row_count = len(rows)

    _auto_size(ws)
    ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)

    # Log export
    await db.execute("""
        INSERT INTO export_history (export_type, filename, filters, row_count)
        VALUES (?, ?, ?, ?)
    """, (export_type, filename, json.dumps(filters), row_count))
    await db.commit()

    return filepath
